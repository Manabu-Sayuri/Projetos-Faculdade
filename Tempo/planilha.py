from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

def salvar_dados_em_planilha(dados, caminho_arquivo="dados_clima.xlsx"):
    # Verifica se a planilha já existe
    if os.path.exists(caminho_arquivo):
        workbook = load_workbook(caminho_arquivo)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        # Cria o cabeçalho na primeira linha
        sheet.append([
            "Data/Hora",
            "Temperatura Máxima",
            "Temperatura Mínima",
            "Umidade Máxima",
            "Umidade Mínima"
        ])

    # Obtém data e hora atual
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Adiciona os dados capturados na planilha
    sheet.append([
        agora,
        dados.get("temperatura_max", ""),
        dados.get("temperatura_min", ""),
        dados.get("umidade_max", ""),
        dados.get("umidade_min", "")
    ])

    # Salva a planilha
    workbook.save(caminho_arquivo)
    print(f"Dados salvos em: {caminho_arquivo}")